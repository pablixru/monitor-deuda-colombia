#!/usr/bin/env python3
"""
Monitor de Deuda Colombia — construcción de los datos del sitio.

Lee los tres libros de Excel oficiales que están en `fuentes/` y escribe los
JSON que consume la página en `public/data/`.

    python scripts/build_data.py

Cada mes basta con reemplazar los archivos de `fuentes/` (conservando los
nombres) y volver a correr el script.

Fuentes
-------
fuentes/tenedores_tes.xlsx          Tenedores de TES clase B      (MinHacienda / Depósito Central de Valores)
fuentes/boletin_deuda_externa.xlsx  Boletín de deuda externa      (Banco de la República)
fuentes/historico_gnc.xlsx          Histórico de deuda del GNC    (MinHacienda — Dirección de Crédito Público)

Salidas
-------
public/data/tes.json      Saldos de TES por tipo de tenedor, mensual, COP millones
public/data/externa.json  Deuda externa bruta del país, mensual, USD millones
public/data/gnc.json      Deuda del Gobierno Nacional Central, mensual
public/data/meta.json     Cortes, TRM, PIB y etiquetas que la página muestra
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
import unicodedata
import warnings
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# La consola de Windows usa cp1252 y revienta al imprimir «→» o «·».
for _flujo in (sys.stdout, sys.stderr):
    try:
        _flujo.reconfigure(encoding="utf-8")
    except Exception:
        pass

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("Falta openpyxl. Instálalo con:  pip install -r requirements.txt")

ROOT = Path(__file__).resolve().parent.parent
FUENTES = ROOT / "fuentes"
SALIDA = ROOT / "public" / "data"

MESES_ES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
            "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
MESES_EN = ["January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"]
# Abreviaturas de mes tal como aparecen en las cabeceras del Banco de la República.
MES_ABBR = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
            "jul": 7, "ago": 8, "sep": 9, "set": 9, "oct": 10, "nov": 11, "dic": 12}


# --------------------------------------------------------------------------
# utilidades
# --------------------------------------------------------------------------

class DatosInesperados(RuntimeError):
    """El Excel no tiene la forma que espera el script."""


def norm(texto) -> str:
    """Minúsculas sin tildes ni espacios repetidos, para comparar etiquetas."""
    if texto is None:
        return ""
    s = unicodedata.normalize("NFKD", str(texto))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def num(v):
    """Devuelve un float si la celda es numérica; si no, None.

    En estos libros los datos faltantes aparecen como celda vacía o como el
    texto '-', y ambos casos deben viajar al JSON como null.
    """
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def redondear(v, decimales: int):
    """Redondeo comercial (medio hacia arriba), conservando None.

    El round() de Python rompe los empates hacia el par, así que 0.85 se
    convertiría en 0.8. Aquí se usa la convención habitual —y la que traían
    los datos publicados— de subir el medio: 0.85 → 0.9.
    """
    if v is None:
        return None
    q = Decimal(1).scaleb(-decimales)
    d = Decimal(repr(float(v))).quantize(q, rounding=ROUND_HALF_UP)
    return int(d) if decimales == 0 else float(d)


def r0(v):
    """Redondeo a entero conservando None."""
    return redondear(v, 0)


def r1(v):
    """Redondeo a 1 decimal conservando None."""
    return redondear(v, 1)


def r3(v):
    """Redondeo a 3 decimales conservando None."""
    return redondear(v, 3)


def ym(fecha: dt.datetime) -> str:
    return f"{fecha.year:04d}-{fecha.month:02d}"


def leer_hoja(wb, nombre: str) -> list[tuple]:
    """Devuelve la hoja completa como lista de tuplas (una por fila)."""
    if nombre not in wb.sheetnames:
        raise DatosInesperados(
            f"El libro no tiene la hoja «{nombre}». Hojas disponibles: {wb.sheetnames}")
    ws = wb[nombre]
    return list(ws.iter_rows(min_row=1, max_row=ws.max_row,
                             max_col=ws.max_column, values_only=True))


def celda(fila: tuple, i: int):
    return fila[i] if i < len(fila) else None


def exigir_etiqueta(filas, idx0: int, esperado: str, hoja: str) -> None:
    """Verifica que la fila `idx0` (0-based) sea la que creemos.

    El formato de estos boletines es estable, así que las filas se leen por
    posición; esta comprobación convierte un cambio de formato en un error
    explícito en vez de en cifras silenciosamente equivocadas.
    """
    fila = filas[idx0] if idx0 < len(filas) else ()
    etiqueta = celda(fila, 0) or celda(fila, 1)
    if esperado not in norm(etiqueta):
        raise DatosInesperados(
            f"[{hoja}] Se esperaba «{esperado}» en la fila {idx0 + 1} y se encontró "
            f"«{etiqueta}». El formato del archivo cambió: revisa scripts/build_data.py.")


def mapa_fechas_columnas(filas, fila_anio: int, fila_mes: int) -> dict[str, int]:
    """Mapa 'YYYY-MM' -> índice de columna, para las hojas del Banco de la República.

    La fila de años sólo trae el año en la primera columna del bloque (y a veces
    con sufijo: '2026 Pr'), y la fila inferior trae el mes de cada columna. Las
    columnas anuales antiguas no tienen mes y por eso quedan fuera.
    """
    anios = filas[fila_anio]
    meses = filas[fila_mes]
    mapa: dict[str, int] = {}
    anio_actual = None
    for i in range(len(anios)):
        crudo = celda(anios, i)
        if crudo is not None:
            m = re.search(r"(19|20)\d{2}", str(crudo))
            if m:
                anio_actual = int(m.group(0))
        mes_txt = norm(celda(meses, i))[:3]
        if anio_actual and mes_txt in MES_ABBR:
            mapa[f"{anio_actual:04d}-{MES_ABBR[mes_txt]:02d}"] = i
    if not mapa:
        raise DatosInesperados(
            "No se pudo leer la cabecera de años/meses; cambió el formato del boletín.")
    return mapa


def mapa_fechas_filas(filas, col: int = 0) -> dict[str, int]:
    """Mapa 'YYYY-MM' -> índice de fila, para las hojas del MinHacienda."""
    mapa: dict[str, int] = {}
    for i, fila in enumerate(filas):
        v = celda(fila, col)
        if isinstance(v, dt.datetime):
            mapa[ym(v)] = i
    return mapa


def serie_por_fila(filas, idx0: int, cols: list[int], f=r1) -> list:
    fila = filas[idx0]
    return [f(num(celda(fila, c))) for c in cols]


def serie_por_columna(filas, col: int, orden: list[str],
                      mapa: dict[str, int], f=r3, escala: float = 1.0) -> list:
    salida = []
    for fecha in orden:
        i = mapa.get(fecha)
        if i is None:
            salida.append(None)
            continue
        v = num(celda(filas[i], col))
        salida.append(f(v * escala) if v is not None else None)
    return salida


# --------------------------------------------------------------------------
# 1) Tenedores de TES
# --------------------------------------------------------------------------

def construir_tes(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(min_row=1, max_row=ws.max_row,
                              max_col=ws.max_column, values_only=True))
    wb.close()

    # La fila de fechas es la única con muchas celdas de tipo fecha.
    fila_fechas = None
    for i, fila in enumerate(filas):
        if sum(1 for c in fila if isinstance(c, dt.datetime)) >= 12:
            fila_fechas = i
            break
    if fila_fechas is None:
        raise DatosInesperados("[TES] No se encontró la fila de fechas.")

    cols = [i for i, c in enumerate(filas[fila_fechas]) if isinstance(c, dt.datetime)]
    fechas = [ym(filas[fila_fechas][i]) for i in cols]

    series: dict[str, list[int]] = {}
    for fila in filas[fila_fechas + 1:]:
        etiqueta = celda(fila, 0)
        if not isinstance(etiqueta, str):
            continue
        etiqueta = etiqueta.strip()
        # La hoja termina con una nota al pie ('*Datos en Millones De Pesos').
        if not etiqueta or etiqueta.startswith("*"):
            continue
        valores = [num(celda(fila, i)) for i in cols]
        if all(v is None for v in valores):
            continue
        series[etiqueta] = [0 if v is None else r0(v) for v in valores]

    if "Total general" not in series:
        raise DatosInesperados("[TES] Falta la fila «Total general».")

    # El total debe cuadrar con la suma de tenedores; si no, hay filas nuevas
    # que la página no estaría mostrando.
    tenedores = [k for k in series if k != "Total general"]
    for j, fecha in enumerate(fechas):
        suma = sum(series[k][j] for k in tenedores)
        total = series["Total general"][j]
        if total and abs(suma - total) / total > 0.001:
            raise DatosInesperados(
                f"[TES] En {fecha} la suma de tenedores ({suma:,.0f}) no cuadra con "
                f"«Total general» ({total:,.0f}).")

    return {"unit": "millones de COP", "dates": fechas, "series": series}


# --------------------------------------------------------------------------
# 2) Deuda externa (Banco de la República)
# --------------------------------------------------------------------------

def construir_externa(path: Path, desde: str = "2001-02") -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    c1 = leer_hoja(wb, "Cuadro 1 Saldo Resumen")
    c3 = leer_hoja(wb, "Cuadro 3 Flujo")
    c4 = leer_hoja(wb, "Cuadro 4 intereses")
    c5 = leer_hoja(wb, "Cuadro 5 DX Publica Monedas")
    c6 = leer_hoja(wb, "Cuadro 6 DX Publica Prestatars")
    c7 = leer_hoja(wb, "Cuadr 7 DX Publica Prestamists")
    wb.close()

    # Cuadro 1 tiene años anuales antiguos a la izquierda y meses desde 2001.
    m1 = mapa_fechas_columnas(c1, 4, 5)
    m3 = mapa_fechas_columnas(c3, 3, 4)
    m4 = mapa_fechas_columnas(c4, 3, 4)
    m5 = mapa_fechas_columnas(c5, 3, 4)
    m6 = mapa_fechas_columnas(c6, 3, 4)
    m7 = mapa_fechas_columnas(c7, 3, 4)

    # Sólo las fechas presentes en todos los cuadros que alimentan las series.
    comunes = sorted(set(m1) & set(m3) & set(m4))
    fechas = [f for f in comunes if f >= desde]
    if not fechas:
        raise DatosInesperados("[Externa] No hay meses en común entre los cuadros.")

    k1 = [m1[f] for f in fechas]
    k3 = [m3[f] for f in fechas]
    k4 = [m4[f] for f in fechas]

    h1 = "Cuadro 1"
    exigir_etiqueta(c1, 6, "total deuda externa bruta", h1)
    exigir_etiqueta(c1, 7, "corto plazo", h1)
    exigir_etiqueta(c1, 9, "sector publico", h1)
    exigir_etiqueta(c1, 12, "sector privado", h1)
    exigir_etiqueta(c1, 15, "porcentaje del pib", h1)

    h3 = "Cuadro 3"
    exigir_etiqueta(c3, 5, "gobierno general", h3)
    exigir_etiqueta(c3, 19, "titulos de deuda", h3)
    exigir_etiqueta(c3, 20, "desembolsos", h3)
    exigir_etiqueta(c3, 21, "amortizaciones", h3)
    exigir_etiqueta(c3, 22, "prestamos", h3)
    exigir_etiqueta(c3, 23, "desembolsos", h3)
    exigir_etiqueta(c3, 24, "amortizaciones", h3)

    h4 = "Cuadro 4"
    exigir_etiqueta(c4, 5, "sector publico", h4)
    exigir_etiqueta(c4, 6, "sector privado", h4)
    exigir_etiqueta(c4, 7, "total", h4)

    # La fila del total del Cuadro 3 flota según el año; se busca por etiqueta.
    fila_total_flujo = next(
        (i for i, fila in enumerate(c3)
         if norm(celda(fila, 0)) == "total deuda externa bruta"), None)
    if fila_total_flujo is None:
        raise DatosInesperados("[Cuadro 3] No se encontró «Total deuda externa bruta».")

    datos = {
        "unit": "millones de USD",
        "dates": fechas,
        "total":   serie_por_fila(c1, 6, k1),
        "cp":      serie_por_fila(c1, 7, k1),
        "lp":      serie_por_fila(c1, 8, k1),
        "pub":     serie_por_fila(c1, 9, k1),
        "pubCp":   serie_por_fila(c1, 10, k1),
        "pubLp":   serie_por_fila(c1, 11, k1),
        "priv":    serie_por_fila(c1, 12, k1),
        "privCp":  serie_por_fila(c1, 13, k1),
        "privLp":  serie_por_fila(c1, 14, k1),
        # El Cuadro 1 publica las razones deuda/PIB como fracción.
        "pibTotal": [r1(v * 100) if v is not None else None
                     for v in serie_por_fila(c1, 16, k1, f=lambda x: x)],
        "pibPub":   [r1(v * 100) if v is not None else None
                     for v in serie_por_fila(c1, 17, k1, f=lambda x: x)],
        "pibPriv":  [r1(v * 100) if v is not None else None
                     for v in serie_por_fila(c1, 18, k1, f=lambda x: x)],
        "intTotal": serie_por_fila(c4, 7, k4),
        "intPub":   serie_por_fila(c4, 5, k4),
        "intPriv":  serie_por_fila(c4, 6, k4),
        "govDesTit": serie_por_fila(c3, 20, k3),
        "govAmoTit": serie_por_fila(c3, 21, k3),
        "govDesPre": serie_por_fila(c3, 23, k3),
        "govAmoPre": serie_por_fila(c3, 24, k3),
        "flujoNeto": serie_por_fila(c3, fila_total_flujo, k3),
        "govNeto":   serie_por_fila(c3, 5, k3),
    }

    # --- foto del último mes disponible en los cuadros 5, 6 y 7 --------------
    corte = max(set(m5) & set(m6) & set(m7))
    i5, i6, i7 = m5[corte], m6[corte], m7[corte]

    def punto(filas, idx0: int, col: int, es: str, en: str):
        v = num(celda(filas[idx0], col))
        return {"es": es, "en": en, "v": r1(v)} if v is not None else None

    monedas = [
        punto(c5, 6, i5, "Dólar (USD)", "US dollar"),
        punto(c5, 7, i5, "Derechos Especiales de Giro", "Special Drawing Rights"),
        punto(c5, 8, i5, "Peso colombiano — TES de no residentes",
              "Colombian peso — non-resident TES"),
        punto(c5, 9, i5, "Euro (EUR)", "Euro (EUR)"),
        punto(c5, 10, i5, "Yen japonés (JPY)", "Japanese yen (JPY)"),
        punto(c5, 11, i5, "Otras monedas", "Other currencies"),
    ]
    acreedores = [
        punto(c7, 6, i7, "Títulos en inversionistas internacionales",
              "Securities held by international investors"),
        punto(c7, 7, i7, "Organismos internacionales", "Multilateral organisations"),
        punto(c7, 12, i7, "Agencias y gobiernos extranjeros",
              "Agencies & foreign governments"),
        punto(c7, 13, i7, "Bancos comerciales y otras inst. financieras",
              "Commercial banks & other financial institutions"),
        punto(c7, 14, i7, "Proveedores", "Suppliers"),
    ]
    multilaterales = [
        punto(c7, 8, i7, "Banco Mundial", "World Bank"),
        punto(c7, 9, i7, "Banco Interamericano de Desarrollo",
              "Inter-American Development Bank"),
        punto(c7, 10, i7, "Fondo Monetario Internacional",
              "International Monetary Fund"),
        punto(c7, 11, i7, "Otros organismos", "Other organisations"),
    ]
    prestatarios = [
        punto(c6, 5, i6, "Gobierno Nacional", "National government"),
        punto(c6, 6, i6, "Entidades descentralizadas nacionales",
              "National decentralised entities"),
        punto(c6, 7, i6, "Municipios y sus entidades", "Municipalities & their entities"),
        punto(c6, 8, i6, "Banco de la República", "Banco de la República"),
        punto(c6, 9, i6, "Departamentos y sus entidades",
              "Departments & their entities"),
    ]

    def limpiar(items, ordenar=True):
        vivos = [x for x in items if x and x["v"]]
        return sorted(vivos, key=lambda x: -x["v"]) if ordenar else vivos

    datos["snapshot"] = {
        "fecha": corte,
        "monedas": limpiar(monedas),
        "acreedores": limpiar(acreedores),
        # Los organismos se dejan en el orden del boletín, no por tamaño.
        "multilaterales": limpiar(multilaterales, ordenar=False),
        "prestatarios": limpiar(prestatarios),
        "pubLargoPlazo": r1(num(celda(c5[12], i5))),
    }
    return datos


# --------------------------------------------------------------------------
# 3) Deuda del Gobierno Nacional Central
# --------------------------------------------------------------------------

def construir_gnc(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    saldos = leer_hoja(wb, "Saldos")
    f_int = leer_hoja(wb, "Fuente - interna")
    f_ext = leer_hoja(wb, "Fuente - externa")
    tasa = leer_hoja(wb, "Tasa - total")
    moneda = leer_hoja(wb, "Moneda - total")
    indic = leer_hoja(wb, "Indicadores")
    perfil = leer_hoja(wb, "Perfil - total")
    wb.close()

    m_saldos = mapa_fechas_filas(saldos)
    fechas = sorted(m_saldos)
    if not fechas:
        raise DatosInesperados("[GNC] La hoja «Saldos» no tiene fechas.")

    def col_saldos(col: int, escala: float, f=r1):
        return serie_por_columna(saldos, col, fechas, m_saldos, f=f, escala=escala)

    m_fint = mapa_fechas_filas(f_int)
    m_fext = mapa_fechas_filas(f_ext)
    m_tasa = mapa_fechas_filas(tasa)
    m_mon = mapa_fechas_filas(moneda)
    m_ind = mapa_fechas_filas(indic)

    def bloque(filas, mapa, cols: dict[str, int]):
        return {k: serie_por_columna(filas, c, fechas, mapa) for k, c in cols.items()}

    datos = {
        "fuente": "Ministerio de Hacienda — Datos históricos de la deuda GNC",
        "corte": fechas[-1],
        "dates": fechas,
        # La hoja publica los saldos en COP millones; la página los usa en billones.
        "copInt": col_saldos(1, 1e-6),
        "copExt": col_saldos(2, 1e-6),
        "copTot": col_saldos(3, 1e-6),
        "usdInt": col_saldos(4, 1.0),
        "usdExt": col_saldos(5, 1.0),
        "usdTot": col_saldos(6, 1.0),
        "fInt": bloque(f_int, m_fint, {
            "tes": 1, "fogafin": 2, "ley546": 3, "reduccion": 4, "agrarios": 5,
            "paz": 6, "seguridad": 7, "deposTesoro": 8, "otros": 9}),
        "fExt": bloque(f_ext, m_fext, {
            "bonos": 1, "bid": 2, "birf": 3, "caf": 4, "otros": 5}),
        "tasa": bloque(tasa, m_tasa, {"fija": 1, "variable": 2}),
        "moneda": bloque(moneda, m_mon, {
            "usd": 1, "eur": 2, "jpy": 3, "chf": 4, "uvr": 5, "cop": 6,
            "local": 7, "extranjera": 8}),
        "ind": bloque(indic, m_ind, {
            "durInt": 1, "durExt": 2, "durTot": 3,
            "vmInt": 4, "vmExt": 5, "vmTot": 6,
            "cupInt": 7, "cupExt": 8, "cupTot": 9}),
        "perfil": construir_perfil(perfil),
    }
    return datos


def construir_perfil(filas) -> dict:
    """Perfil de vencimientos del último corte, en billones de COP.

    La hoja repite, para cada corte, un bloque de tres filas
    (Amortizaciones / Intereses / Total); la fecha va en la fila del medio.
    """
    fila_anios = None
    for i, fila in enumerate(filas[:40]):
        if norm(celda(fila, 1)) in ("periodo de servicio", "período de servicio"):
            fila_anios = i
            break
    if fila_anios is None:
        raise DatosInesperados("[GNC] No se encontró la cabecera del perfil de vencimientos.")

    cols, anios = [], []
    for i in range(2, len(filas[fila_anios])):
        v = celda(filas[fila_anios], i)
        if isinstance(v, (int, float)) and 1990 < float(v) < 2200:
            cols.append(i)
            anios.append(int(v))

    ultimo = max((i for i, f in enumerate(filas) if isinstance(celda(f, 0), dt.datetime)),
                 default=None)
    if ultimo is None:
        raise DatosInesperados("[GNC] El perfil de vencimientos no tiene fechas.")

    # La fecha marca la fila de 'Intereses': una arriba y una abajo cierran el bloque.
    etiquetas = {"amortizaciones": "Amortizaciones", "intereses": "Intereses", "total": "Total"}
    series = []
    for idx in (ultimo - 1, ultimo, ultimo + 1):
        if idx < 0 or idx >= len(filas):
            continue
        etiqueta = norm(celda(filas[idx], 1))
        if etiqueta not in etiquetas:
            raise DatosInesperados(
                f"[GNC] Perfil: se esperaba Amortizaciones/Intereses/Total y se "
                f"encontró «{celda(filas[idx], 1)}» en la fila {idx + 1}.")
        vals = []
        for c in cols:
            v = num(celda(filas[idx], c))
            vals.append(None if v is None else r1(v * 1e-6))
        series.append({"label": etiquetas[etiqueta], "vals": vals})

    return {"anios": anios, "series": series}


# --------------------------------------------------------------------------
# 4) meta.json — cortes y parámetros que la página muestra en texto
# --------------------------------------------------------------------------

def construir_meta(tes: dict, externa: dict, gnc: dict, trm: float, pib: float) -> dict:
    def etiquetas(fecha: str) -> dict:
        anio, mes = fecha.split("-")
        i = int(mes) - 1
        return {
            "iso": fecha,
            "es": f"{MESES_ES[i]} de {anio}",
            "en": f"{MESES_EN[i]} {anio}",
            "esCorto": f"{MESES_ES[i].capitalize()} {anio}",
            "enCorto": f"{MESES_EN[i]} {anio}",
            "anio": anio,
        }

    return {
        "generado": dt.datetime.now().strftime("%Y-%m-%d"),
        "trm": trm,
        "pibCop": pib,
        "cortes": {
            "tes": etiquetas(tes["dates"][-1]),
            "externa": etiquetas(externa["dates"][-1]),
            "gnc": etiquetas(gnc["dates"][-1]),
        },
        "rangos": {
            "tes": {"desde": tes["dates"][0][:4], "hasta": tes["dates"][-1][:4]},
            "externa": {"desde": externa["dates"][0][:4], "hasta": externa["dates"][-1][:4]},
            "gnc": {"desde": gnc["dates"][0][:4], "hasta": gnc["dates"][-1][:4]},
        },
    }


# --------------------------------------------------------------------------

def trm_del_corte(corte: str):
    """TRM del último día hábil del mes de corte, desde Datos Abiertos Colombia.

    La TRM sólo existe para días hábiles, así que no se pide la fecha exacta de
    fin de mes sino la última disponible hasta ese día. Devuelve None si no se
    puede consultar: en ese caso se conserva la del mes anterior.
    """
    import json as _json
    import urllib.parse
    import urllib.request

    anio, mes = (int(x) for x in corte.split("-"))
    ultimo = dt.date(anio + (mes == 12), mes % 12 + 1, 1) - dt.timedelta(days=1)
    consulta = urllib.parse.urlencode({
        "$where": f"vigenciadesde <= '{ultimo.isoformat()}T23:59:59'",
        "$order": "vigenciadesde DESC",
        "$limit": 1,
    })
    url = f"https://www.datos.gov.co/resource/mcec-87by.json?{consulta}"
    try:
        with urllib.request.urlopen(url, timeout=40) as r:
            filas = _json.loads(r.read().decode("utf-8"))
        if not filas:
            return None
        valor = float(filas[0]["valor"])
        fecha = filas[0]["vigenciadesde"][:10]
        print(f"  · TRM del {fecha}: {valor:,.2f} (Datos Abiertos Colombia)")
        return valor
    except Exception as e:
        print(f"  · No se pudo consultar la TRM ({e}); se conserva la anterior.")
        return None


def escribir(destino: Path, datos: dict) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)
    with destino.open("w", encoding="utf-8") as fh:
        json.dump(datos, fh, ensure_ascii=False, separators=(",", ":"))
    print(f"  ✓ {destino.relative_to(ROOT).as_posix():<26} {destino.stat().st_size / 1024:7.1f} kB")


def main() -> int:
    ap = argparse.ArgumentParser(description="Construye los datos del Monitor de Deuda.")
    ap.add_argument("--tes", type=Path, default=FUENTES / "tenedores_tes.xlsx")
    ap.add_argument("--externa", type=Path, default=FUENTES / "boletin_deuda_externa.xlsx")
    ap.add_argument("--gnc", type=Path, default=FUENTES / "historico_gnc.xlsx")
    ap.add_argument("--salida", type=Path, default=SALIDA)
    ap.add_argument("--trm", default=None,
                    help="TRM del corte del boletín externo, o «auto» para consultarla.")
    ap.add_argument("--pib", type=float, default=None,
                    help="PIB nominal en billones de COP.")
    args = ap.parse_args()

    faltan = [p for p in (args.tes, args.externa, args.gnc) if not p.exists()]
    if faltan:
        print("No se encontraron estos archivos:", file=sys.stderr)
        for p in faltan:
            print(f"  · {p}", file=sys.stderr)
        return 1

    print("Leyendo fuentes…")
    tes = construir_tes(args.tes)
    print(f"  · TES      {tes['dates'][0]} → {tes['dates'][-1]}  "
          f"({len(tes['dates'])} meses, {len(tes['series']) - 1} tipos de tenedor)")
    externa = construir_externa(args.externa)
    print(f"  · Externa  {externa['dates'][0]} → {externa['dates'][-1]}  "
          f"({len(externa['dates'])} meses, foto a {externa['snapshot']['fecha']})")
    gnc = construir_gnc(args.gnc)
    print(f"  · GNC      {gnc['dates'][0]} → {gnc['dates'][-1]}  ({len(gnc['dates'])} meses)")

    # La TRM y el PIB no vienen en los libros: se conservan los del mes anterior
    # salvo que se pasen por línea de comandos.
    meta_previo = {}
    ruta_meta = args.salida / "meta.json"
    if ruta_meta.exists():
        meta_previo = json.loads(ruta_meta.read_text(encoding="utf-8"))
    anterior = meta_previo.get("trm", 3621.86)
    if args.trm is None:
        trm = anterior
    elif str(args.trm).lower() == "auto":
        trm = trm_del_corte(externa["dates"][-1]) or anterior
    else:
        trm = float(args.trm)
    pib = args.pib if args.pib is not None else meta_previo.get("pibCop", 1928)

    print("\nEscribiendo datos…")
    escribir(args.salida / "tes.json", tes)
    escribir(args.salida / "externa.json", externa)
    escribir(args.salida / "gnc.json", gnc)
    escribir(args.salida / "meta.json", construir_meta(tes, externa, gnc, trm, pib))

    print(f"\nListo. Corte TES/GNC: {gnc['corte']} · deuda externa: "
          f"{externa['dates'][-1]} · TRM {trm:,.2f} · PIB ${pib:,.0f} B")
    if args.trm is None or args.pib is None:
        print("Recuerda revisar --trm y --pib si cambiaron este mes (--trm auto los consulta).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
